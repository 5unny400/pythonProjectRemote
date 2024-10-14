"""
@FileName：test
@Description: 将{service}-requirements.txt文件中的依赖版本与依赖版本维护.xlsx文件中的依赖版本进行比对，输出差异
@Author：shenxinyuan
@Time：2024/10/10
"""
# 读取依赖版本维护.xlsx文件的第一个 sheet 中的前两列数据
import openpyxl

# 打开文件
wb = openpyxl.load_workbook('依赖版本维护.xlsx')
# key 为依赖名称，value 为推荐版本
standard_version = {}
for sheetname in wb.sheetnames:
    ws = wb[sheetname]
    # 读取前两列数据，不读取第一行
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        if row[0].value:
            # 格式化打印数据
            # print(f'{row[0].value:<20} {row[1].value}')
            standard_version[row[0].value.lower()] = row[1].value
# 关闭文件
wb.close()


for service in ['bfa', 'cbl']:
    print('--------------------------------------------------------')
    print(f'------------------{service}依赖比对结果-------------------------')
    with open(f'{service}-requirements.txt', 'r', encoding='utf-8') as f:
        index = 0
        # 读取前两列数据，不读取第一行，列与列之间用空格分隔
        for line in f.readlines()[1:]:
            name, version = line.strip().split()
            name = name.strip()
            if name.lower() not in standard_version:
                index += 1
                print(f'{index:<3}{service:<5} {name:<20}推荐依赖不存在，实际版本为 {version}')
            elif name.lower() in standard_version and version != standard_version[name.lower()]:
                index += 1
                print(f'{index:<3}{service:<5} {name:<20}推荐版本为 {standard_version[name.lower()]}，实际版本为 {version}')
        # 关闭文件
        f.close()
